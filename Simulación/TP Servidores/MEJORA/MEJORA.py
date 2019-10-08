import numpy as np
from openpyxl import Workbook

RELOJ = 0.0
ListaArribos = []
ListaPartidas = []
#Gráficas
RelojEnT = []
NroClientesCola = []
UtiServ = []
DemoraProm = []
#Excel
wb = Workbook()
ws = wb.active


class Simulacion():

    def __init__(self, servidor, tpoEntreArribos, tpoDeServicio):
        self.nroServidor = servidor
        self.tm_entre_arribos = tpoEntreArribos
        self.tm_servicio = tpoDeServicio
        self.ts_acumulado = 0.0
        self.demora_acumulada = 0.0
        self.area_q_t = 0.0
        self.tiempo_ultimo_evento = 0.0
        self.nro_clientes_cola = 0
        self.completaron_demora = 0
        self.iniciado = False
        self.estado_servidor = ""  # D - disponible | O - ocupado
        self.proximo_evento = ""  # A - Arribo | P - partida
        self.cola = []
        self.cola.append(0)
        # las listas siguientes se utilizan sólo con el fin de generar las gráficas
        self.cliColaEnT = []
        self.relojEnT = []
        self.tsAcuEnT = []

        self.estado_servidor = "D"
        self.proximo_evento = ""
        self.iniciado = False

    def tiempos(self, evento, letra):
        global RELOJ
        if letra == 'A':
            self.tiempo_ultimo_evento = RELOJ
            RELOJ = evento[0]
        else:
            RELOJ = evento[0]
        

    def arribo(self, evento):
        global RELOJ
        global ListaArribos

        indice = evento[1] - 1  # posicion en lista arribos

        if self.nroServidor < 3:
            ListaArribos[indice] = [RELOJ + np.random.exponential(self.tm_entre_arribos), evento[1]]

            if self.estado_servidor == "D":
                self.estado_servidor = "O"
                ListaPartidas[indice] = [RELOJ + np.random.exponential(self.tm_servicio), evento[1]]
                self.ts_acumulado += (ListaPartidas[indice][0] - RELOJ)
                self.completaron_demora += 1
                # grafica
                """
                sim1.relojEnT.append(RELOJ)  
                sim1.cliColaEnT.append(sim1.nro_clientes_cola)
                sim1.tsAcuEnT.append(sim1.ts_acumulado)
                """
            else:
                self.area_q_t += len(self.cola) * (RELOJ - self.tiempo_ultimo_evento)
                self.nro_clientes_cola += 1
                self.cola.append(RELOJ)
                # grafica
                """
                sim1.cliColaEnT.append(sim1.nro_clientes_cola)
                sim1.relojEnT.append(RELOJ)
                sim1.tsAcuEnT.append(sim1.ts_acumulado)
                """
        else:
            if self.nroServidor == 3:
                if self.estado_servidor == 'D':
                    self.estado_servidor = "O"
                    ListaPartidas[indice] = [RELOJ + np.random.exponential(self.tm_servicio), evento[1]]
                    self.ts_acumulado += (ListaPartidas[indice][0] - RELOJ)
                    self.completaron_demora += 1
                else:
                    self.area_q_t += len(self.cola) * (RELOJ - self.tiempo_ultimo_evento)
                    self.nro_clientes_cola += 1
                    self.cola.append(RELOJ)

            elif self.nroServidor == 4:
                if self.estado_servidor == 'D':
                    self.estado_servidor = "O"
                    ListaPartidas[indice] = [RELOJ + np.random.exponential(self.tm_servicio), evento[1]]
                    self.ts_acumulado += (ListaPartidas[indice][0] - RELOJ)
                    self.completaron_demora += 1
                else:
                    self.area_q_t += len(self.cola) * (RELOJ - self.tiempo_ultimo_evento)
                    self.nro_clientes_cola += 1
                    self.cola.append(RELOJ)
                
            else:
                if self.estado_servidor == 'D':
                    self.estado_servidor = "O"
                    ListaPartidas[indice] = [RELOJ + np.random.exponential(self.tm_servicio), evento[1]]
                    self.ts_acumulado += (ListaPartidas[indice][0] - RELOJ)
                    self.completaron_demora += 1
                else:
                    self.area_q_t += len(self.cola) * (RELOJ - self.tiempo_ultimo_evento)
                    self.nro_clientes_cola += 1
                    self.cola.append(RELOJ)

    def partida(self, evento):
        global RELOJ
        global ListaPartidas
        indice = evento[1] - 1  # posicion en lista partida

        if self.nroServidor < 3:
            if len(self.cola) > 0:
                ListaPartidas[indice] = [RELOJ + np.random.exponential(self.tm_servicio), evento[1]]
                self.demora_acumulada += RELOJ - self.cola[0]
                self.completaron_demora += 1
                self.ts_acumulado += (ListaPartidas[indice][0] - RELOJ)
                self.area_q_t += (len(self.cola) * (RELOJ - self.tiempo_ultimo_evento))
                self.nro_clientes_cola -= 1
                self.cola.pop(0)
                # grafica
                """
                sim1.cliColaEnT.append(sim1.nro_clientes_cola)
                sim1.relojEnT.append(RELOJ)
                sim1.tsAcuEnT.append(sim1.ts_acumulado)
                """
            else:
                self.estado_servidor = "D"
                ListaPartidas[indice] = [999999, evento[1]]
                # grafica
                """
                sim1.relojEnT.append(RELOJ)
                sim1.cliColaEnT.append(sim1.nro_clientes_cola)
                sim1.tsAcuEnT.append(sim1.ts_acumulado)
                """
        else:
            if self.nroServidor == 3:
                if len(self.cola) > 0:
                    ListaPartidas[indice] = [RELOJ + np.random.exponential(self.tm_servicio), evento[1]]
                    self.area_q_t += (len(self.cola) * (RELOJ - self.tiempo_ultimo_evento))
                    self.demora_acumulada += (RELOJ - self.cola[0])
                    self.completaron_demora += 1
                    
                    self.cola.pop(0)
                    if len(self.cola) == 0:
                        self.estado_servidor ='D'
                    
                else:
                    self.estado_servidor = 'D'
                    ListaPartidas[indice] = [999999, evento[1]]
            elif self.nroServidor == 4:
                if len(self.cola) > 0:
                    ListaPartidas[indice] = [RELOJ + np.random.exponential(self.tm_servicio), evento[1]]
                    self.area_q_t += (len(self.cola) * (RELOJ - self.tiempo_ultimo_evento))
                    self.demora_acumulada += (RELOJ - self.cola[0])
                    self.completaron_demora += 1
                    
                    self.cola.pop(0)
                    if len(self.cola) == 0:
                        self.estado_servidor ='D'
                    
                else:
                    self.estado_servidor = 'D'
                    ListaPartidas[indice] = [999999, evento[1]]
            else:
                if len(self.cola) > 0:
                    ListaPartidas[indice] = [RELOJ + np.random.exponential(self.tm_servicio), evento[1]]
                    self.area_q_t += (len(self.cola) * (RELOJ - self.tiempo_ultimo_evento))
                    self.demora_acumulada += (RELOJ - self.cola[0])
                    self.completaron_demora += 1
                    
                    self.cola.pop(0)
                    if len(self.cola) == 0:
                        self.estado_servidor ='D'
                    
                else:
                    self.estado_servidor = 'D'
                    ListaPartidas[indice] = [999999, evento[1]]


def run(server1, server2, server3, server4, server5):
    global ListaArribos
    global ListaPartidas
    global RELOJ
    global ColaUnica
    print("Inicializando simulacion")

    # Tiempo del primer arribo
    ListaArribos.append([np.random.exponential(server1.tm_entre_arribos), 1])
    ListaArribos.append([np.random.exponential(server2.tm_entre_arribos), 2])
    ListaArribos.append([99999, 3])
    ListaArribos.append([99999, 4])
    ListaArribos.append([99999, 5])

    # Numero grande para asegurar que el primer evento sea un arribo
    ListaPartidas.append([999999, 1])
    ListaPartidas.append([999999, 2])
    ListaPartidas.append([999999, 3])
    ListaPartidas.append([999999, 4])
    ListaPartidas.append([999999, 5])

    while RELOJ < 2000:
        minArribo = min(ListaArribos)
        minPartida = min(ListaPartidas)

        if minArribo < minPartida:
            print("reloj: " + str(round(RELOJ,2)))
            RelojEnT.append(round(RELOJ,2))
            nroServer = minArribo[1]
            if nroServer == 1:
                server1.tiempos(minArribo, 'A')
                server1.arribo(minArribo)
            else:
                server2.tiempos(minArribo, 'A')
                server2.arribo(minArribo)

        else:
            print("reloj: " + str(round(RELOJ,2)))
            RelojEnT.append(round(RELOJ,2))
            nroServer = minPartida[1]
            if nroServer == 1:
                server1.tiempos(minPartida, 'P')
                server1.partida(minPartida)
                if server3.estado_servidor == 'O' and server4.estado_servidor == 'O' and server5.estado_servidor == 'O':
                    long_colas = [server3.nro_clientes_cola, server4.nro_clientes_cola, server5.nro_clientes_cola]
                    indice = long_colas.index(min(long_colas)) + 3
                    if indice == 3:
                        server3.tiempos(minPartida, 'A')
                        server3.arribo(minPartida)
                    elif indice == 4:
                        server4.tiempos(minPartida, 'A')
                        server4.arribo(minPartida)
                    elif indice == 5:
                        server5.tiempos(minPartida, 'A')
                        server5.arribo(minPartida)
                elif server3.estado_servidor == 'D':
                    minPartida[1] = minPartida[1] + 2
                    server3.tiempos(minPartida, 'A')
                    server3.arribo(minPartida)
                elif server4.estado_servidor == 'D':
                    minPartida[1] = minPartida[1] + 3
                    server4.tiempos(minPartida, 'A')
                    server4.arribo(minPartida)
                else:
                    minPartida[1] = minPartida[1] + 4
                    server5.tiempos(minPartida, 'A')
                    server5.arribo(minPartida)
            elif nroServer == 2:
                server2.tiempos(minPartida, 'P')
                server2.partida(minPartida)
                if server3.estado_servidor == 'O' and server4.estado_servidor == 'O' and server5.estado_servidor == 'O':
                    long_colas = [server3.nro_clientes_cola, server4.nro_clientes_cola, server5.nro_clientes_cola]
                    indice = long_colas.index(min(long_colas)) + 3
                    if indice == 3:
                        server3.tiempos(minPartida, 'A')
                        server3.arribo(minPartida)
                    elif indice == 4:
                        server4.tiempos(minPartida, 'A')
                        server4.arribo(minPartida)
                    elif indice == 5:
                        server5.tiempos(minPartida, 'A')
                        server5.arribo(minPartida)
                elif server3.estado_servidor == 'D':
                    minPartida[1] = minPartida[1] + 1
                    server3.tiempos(minPartida, 'A')
                    server3.arribo(minPartida)
                elif server4.estado_servidor == 'D':
                    minPartida[1] = minPartida[1] + 2
                    server4.tiempos(minPartida, 'A')
                    server4.arribo(minPartida)
                else:
                    minPartida[1] = minPartida[1] + 3
                    server5.tiempos(minPartida, 'A')
                    server5.arribo(minPartida)
            elif nroServer == 3:
                server3.tiempos(minPartida, 'P')
                server3.partida(minPartida)
            elif nroServer == 4:
                server4.tiempos(minPartida, 'P')
                server4.partida(minPartida)
            elif nroServer == 5:
                server5.tiempos(minPartida, 'P')
                server5.partida(minPartida)
        NroClientesCola.append(nro_prom_clientes_cola(server1, server2, server3, server4, server5))
        UtiServ.append(utilizacion_prom_servidor(server1, server2, server3, server4, server5))
        DemoraProm.append(demora_prom_cliente(server1, server2, server3, server4, server5))
    return server1, server2, server3, server4, server5


def nro_prom_clientes_cola(server1, server2, server3, server4, server5):
    global RELOJ
    if RELOJ != 0:
        nro_clientes_prom_cola1 = server1.area_q_t / RELOJ
        nro_clientes_prom_cola2 = server2.area_q_t / RELOJ
        nro_clientes_prom_cola3 = server3.area_q_t / RELOJ
        nro_clientes_prom_cola4 = server4.area_q_t / RELOJ
        nro_clientes_prom_cola5 = server5.area_q_t / RELOJ
        return round(nro_clientes_prom_cola1,2), round(nro_clientes_prom_cola2,2), round(nro_clientes_prom_cola3,2), round(nro_clientes_prom_cola4,2), round(nro_clientes_prom_cola5,2)
    else:
        nro_clientes_prom_cola = 0  # no tiene sentido, ver
        return nro_clientes_prom_cola


def utilizacion_prom_servidor(server1, server2, server3, server4, server5):
    if RELOJ != 0:
        utilizacion_prom_servidor1 = server1.ts_acumulado / RELOJ
        if (utilizacion_prom_servidor1 > 100.00):
            utilizacion_prom_servidor1 = 100.00
        
        utilizacion_prom_servidor2 = server2.ts_acumulado / RELOJ
        if (utilizacion_prom_servidor2 > 100.00):
            utilizacion_prom_servidor2 = 100.00
        
        utilizacion_prom_servidor3 = server3.ts_acumulado / RELOJ
        if (utilizacion_prom_servidor3 > 100.00):
            utilizacion_prom_servidor3 = 100.00
        
        utilizacion_prom_servidor4 = server4.ts_acumulado / RELOJ
        if (utilizacion_prom_servidor4>  100.00):
            utilizacion_prom_servidor4 = 100.00
        
        utilizacion_prom_servidor5 = server5.ts_acumulado / RELOJ
        if (utilizacion_prom_servidor5 > 100.00):
            utilizacion_prom_servidor5 = 100.00
        return round((utilizacion_prom_servidor1 * 100),2), round((utilizacion_prom_servidor2 * 100),2), round((utilizacion_prom_servidor3 * 100),2), round((utilizacion_prom_servidor4 * 100),2), round((utilizacion_prom_servidor5 * 100),2)
    else:
        utilizacion_prom_servidor = 0  # no tiene sentido, ver
        return utilizacion_prom_servidor


def demora_prom_cliente(server1, server2, server3, server4, server5):
    if server1.completaron_demora > 0:
        demora_prom_cliente1 = server1.demora_acumulada / server1.completaron_demora
    else:
        demora_prom_cliente1 = 0
    if server2.completaron_demora > 0:
        demora_prom_cliente2 = server2.demora_acumulada / server2.completaron_demora
    else:
        demora_prom_cliente2 = 0
    if server3.completaron_demora > 0:
        demora_prom_cliente3 = server3.demora_acumulada / server3.completaron_demora
    else:
        demora_prom_cliente3 = 0
    if server4.completaron_demora > 0:
        demora_prom_cliente4 = server4.demora_acumulada / server4.completaron_demora
    else:
        demora_prom_cliente4 = 0
    if server5.completaron_demora > 0:
        demora_prom_cliente5 = server5.demora_acumulada / server5.completaron_demora
    else:
        demora_prom_cliente5 = 0

    demoratotal = (demora_prom_cliente1+demora_prom_cliente2+demora_prom_cliente3+demora_prom_cliente4+demora_prom_cliente5)/5
    return round(demoratotal,2)



def reportes(server1, server2, server3, server4, server5):
    print(" Estadísticos ".center(101, '='), '\n')
    print("Cantidad promedio de clientes en cola: ", '\n')
    res = nro_prom_clientes_cola(server1, server2, server3, server4, server5)
    print("Cola server 1: ", str(res[0]))
    print("Cola server 2: ", str(res[1]))
    print("Cola server 3: ", str(res[2]))
    print("Cola server 4: ", str(res[3]))
    print("Cola server 5: ", str(res[4]))
    print('\n')

    print("Utilización promedio de los servers: ", '\n')
    res = utilizacion_prom_servidor(server1, server2, server3, server4, server5)
    print("Server 1: ", str(round(res[0], 2)), '%')
    print("Server 2: ", str(round(res[1], 2)), '%')
    print("Server 3: ", str(round(res[2], 2)), '%')
    print("Server 4: ", str(round(res[3], 2)), '%')
    print("Server 5: ", str(round(res[4], 2)), '%')
    print('\n')

    res = demora_prom_cliente(server1, server2, server3, server4, server5)
    print("Promedio del total de demora en los servidores: ", str(res))

    print(" \/\/\/\/\/ ".center(101, '='))
    #graficas 
    #graficaPromCliEnCola()
    #graficaUtiServidor()


if __name__ == '__main__':
    server1 = Simulacion(servidor=1, tpoEntreArribos=10.0, tpoDeServicio=7.0)
    server2 = Simulacion(servidor=2, tpoEntreArribos=10.0, tpoDeServicio=5.0)
    server3 = Simulacion(servidor=3, tpoEntreArribos=10.0, tpoDeServicio=6.0)
    server4 = Simulacion(servidor=4, tpoEntreArribos=10.0, tpoDeServicio=5.0)
    server5 = Simulacion(servidor=5, tpoEntreArribos=10.0, tpoDeServicio=5.0)

    run(server1, server2, server3, server4, server5)
    reportes(server1, server2, server3, server4, server5)
    #Excel
    for i in range(len(RelojEnT)):
        ws.cell(column=1, row=i+1).value = RelojEnT[i]
    for i in range(len(DemoraProm)):
        ws.cell(column=2, row=i+1).value = DemoraProm[i]
    for i in range(len(NroClientesCola)):
        ws.cell(column=3, row=i+1).value = NroClientesCola[i][0]
        ws.cell(column=4, row=i+1).value = NroClientesCola[i][1]
        ws.cell(column=5, row=i+1).value = NroClientesCola[i][2]
        ws.cell(column=6, row=i + 1).value = NroClientesCola[i][3]
        ws.cell(column=7, row=i + 1).value = NroClientesCola[i][4]
    for i in range(len(UtiServ)):
        ws.cell(column=8, row=i+1).value = UtiServ[i][0]
        ws.cell(column=9, row=i+1).value = UtiServ[i][1]
        ws.cell(column=10, row=i+1).value = UtiServ[i][2]
        ws.cell(column=11, row=i+1).value = UtiServ[i][3]
        ws.cell(column=12,row=i+1).value = UtiServ[i][4]

    wb.save("resultadosMEJORAS.xlsx")
